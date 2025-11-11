import streamlit as st
import openpyxl
import io
import random
import logging
import numpy as np
import pandas as pd
import altair as alt
from scipy.signal import medfilt
from scipy.ndimage import gaussian_filter1d

# --- Configuración de Logging ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# --- Constantes Globales ---
ARCHIVOS_CON_LIMITE = ["1. OQ_MAPEO.xlsm", "4. PQ_RUTA_20.xlsm", "5. PQ_RUTA_80.xlsm"]
HOJAS_A_IGNORAR = ["CONSOLIDADO", "GRAFICOS", "RESUMEN", "TABLA", "RESULTADOS", "SUMMARY", "GRAFICO"] 

# --- CONFIGURACIÓN BASE (Presets V23) ---
# Añadimos 'aplanado_global' (default 0.0 = sin efecto)
CONFIGURACION_BASE = {
    "OQ Mapeo": {
        "archivo": "1. OQ MAPEO 72 INV.xlsm",        
        "variacion_min": 0.01, "variacion_max": 0.02,
        "amplitud": 0.30, "sigma": 12, "punto_pico": 0.5,
        "offset_min": -0.5, "offset_max": 0.0,
        "escala_picos": 0.8,
        "aplanado_global": 0.0
    },
    "OQ Apertura": {
        "archivo": "2. OQ APERTURA 72 INV.xlsm",     
        "variacion_min": 0.03, "variacion_max": 0.05,
        "amplitud": 0.40, "sigma": 8, "punto_pico": 0.6,
        "offset_min": -1.0, "offset_max": -0.2,
        "escala_picos": 0.7,
        "aplanado_global": 0.0
    },
    "OQ Apagado": {
        "archivo": "3. OQ APAGADO 72 INV.xlsm",      
        "variacion_min": 0.01, "variacion_max": 0.02,
        "amplitud": 0.50, "sigma": 20, "punto_pico": 0.4,
        "offset_min": -1.2, "offset_max": -0.3,
        "escala_picos": 0.9,
        "aplanado_global": 0.0
    },
    "PQ Ruta 20%": {
        "archivo": "4. PQ RUTA 20 72 INV.xlsm",      
        "variacion_min": 0.02, "variacion_max": 0.04,
        "amplitud": 0.35, "sigma": 12, "punto_pico": 0.5,
        "offset_min": -0.9, "offset_max": -0.2,
        "escala_picos": 0.8,
        "aplanado_global": 0.0
    },
    "PQ Ruta 80%": {
        "archivo": "5. PQ RUTA 80 72 INV.xlsm",      
        "variacion_min": 0.02, "variacion_max": 0.04,
        "amplitud": 0.35, "sigma": 12, "punto_pico": 0.5,
        "offset_min": -0.9, "offset_max": -0.2,
        "escala_picos": 0.8,
        "aplanado_global": 0.0
    },
    "PQ Apertura 20%": {
        "archivo": "6. PQ APERTURA 20 72 INV.xlsm",  
        "variacion_min": 0.03, "variacion_max": 0.05,
        "amplitud": 0.40, "sigma": 8, "punto_pico": 0.6,
        "offset_min": -1.0, "offset_max": -0.2,
        "escala_picos": 0.7,
        "aplanado_global": 0.0
    },
    "PQ Apertura 80%": {
        "archivo": "7. PQ APERTURA 80 72 INV.xlsm",  
        "variacion_min": 0.03, "variacion_max": 0.05,
        "amplitud": 0.40, "sigma": 8, "punto_pico": 0.6,
        "offset_min": -1.0, "offset_max": -0.2,
        "escala_picos": 0.7,
        "aplanado_global": 0.0
    },
    "PQ Apagado 20%": {
        "archivo": "8. PQ APAGADO 20 72 INV.xlsm",   
        "variacion_min": 0.01, "variacion_max": 0.02,
        "amplitud": 0.50, "sigma": 20, "punto_pico": 0.4,
        "offset_min": -1.2, "offset_max": -0.3,
        "escala_picos": 0.9,
        "aplanado_global": 0.0
    },
    "PQ Apagado 80%": {
        "archivo": "9. PQ APAGADO 80 72 INV.xlsm",   
        "variacion_min": 0.01, "variacion_max": 0.02,
        "amplitud": 0.50, "sigma": 20, "punto_pico": 0.4,
        "offset_min": -1.2, "offset_max": -0.3,
        "escala_picos": 0.9,
        "aplanado_global": 0.0
    },
}

# --- FUNCIONES DE GENERACIÓN DE CURVAS (El "Motor") ---

@st.cache_data(show_spinner=False)
def generar_deriva_gaussiana(longitud, amplitud, sigma, seed):
    """(PASO 3) Genera una curva de deriva suave (aditiva) única por DL."""
    np.random.seed(seed)
    try:
        ruido_base = np.random.randn(longitud)
        deriva_suave = gaussian_filter1d(ruido_base, sigma=sigma)
        max_abs = np.max(np.abs(deriva_suave))
        if max_abs > 1e-6: deriva_normalizada = deriva_suave / max_abs
        else: deriva_normalizada = np.zeros(longitud)
        deriva_final = deriva_normalizada * amplitud
        fade_len = min(longitud // 10, int(sigma * 3))
        if fade_len > 1:
            fade_in = np.linspace(0, 1, fade_len)
            deriva_final[:fade_len] *= fade_in
            fade_out = np.linspace(1, 0, fade_len)
            deriva_final[-fade_len:] *= fade_out
        return deriva_final
    except Exception: return np.zeros(longitud)

@st.cache_data(show_spinner=False)
def generar_curva_multiplicativa(longitud, variacion_percent, punto_pico_frac):
    """(PASO 2) Genera una curva de multiplicación que vuelve a 1.0."""
    try:
        factor_max = 1.0 + variacion_percent
        punto_pico_idx = int(longitud * punto_pico_frac)
        if punto_pico_idx <= 0: punto_pico_idx = 1
        if punto_pico_idx >= longitud: punto_pico_idx = longitud - 1
        fase_subida = np.linspace(1.0, factor_max, punto_pico_idx)
        fase_bajada = np.linspace(factor_max, 1.0, longitud - punto_pico_idx)
        curva_multi = np.concatenate((fase_subida[:-1], fase_bajada))
        if len(curva_multi) != longitud:
            x_original = np.linspace(0, 1, len(curva_multi))
            x_nuevo = np.linspace(0, 1, longitud)
            curva_multi = np.interp(x_nuevo, x_original, curva_multi)
        return curva_multi
    except Exception: return np.ones(longitud)

#@st.cache_data(show_spinner=False)
def aplicar_pipeline_a_columna(datos_np, config_dl, seed):
    """
    Aplica el pipeline base (Pasos 1-4) a una columna.
    El aplanado global se aplica DESPUÉS de esta función.
    """
    longitud_actual = len(datos_np)
    if longitud_actual < 20:
        return datos_np

    col_seed = seed + hash(config_dl['dl_nombre']) % 1000
    random.seed(col_seed)
    np.random.seed(col_seed)

    # PASO 1: Lógica de Escala de Picos (Editor Individual)
    datos_base_limpios = medfilt(datos_np, kernel_size=3)
    picos_originales = datos_np - datos_base_limpios
    picos_escalados = picos_originales * config_dl["escala_picos"]
    datos_base = datos_base_limpios + picos_escalados

    # PASO 2: EXTRAPOLACIÓN
    curva_multi_dl = generar_curva_multiplicativa(longitud_actual, config_dl["variacion_percent"], config_dl["punto_pico_frac"])
    datos_extrapolados = datos_base * curva_multi_dl
    
    # PASO 3: DERIVA DE REALISMO
    deriva = generar_deriva_gaussiana(longitud_actual, config_dl["amplitud"], config_dl["sigma"], seed=col_seed + 1)
    datos_con_deriva = datos_extrapolados + deriva
    
    # PASO 4: OFFSET BASE
    datos_finales = datos_con_deriva + config_dl["offset_base"]
    
    return datos_finales

# --- Lógica de Edición de Sección (V22) ---
def aplicar_edicion_seccion(datos_originales, inicio_idx, fin_idx, ajuste_offset, ajuste_aplanado):
    """
    Aplica edición a una sección específica de los datos.
    'ajuste_aplanado' (0.0 a 1.0) controla la mezcla con una línea recta (estirar hilo).
    """
    datos_editados = datos_originales.copy()
    
    inicio_idx = max(0, min(inicio_idx, len(datos_originales)-1))
    fin_idx = max(0, min(fin_idx, len(datos_originales)-1))
    
    if inicio_idx >= fin_idx or (fin_idx - inicio_idx) < 1:
        return datos_originales
    
    seccion = datos_originales[inicio_idx:fin_idx+1]
    
    valor_inicio = seccion[0]
    valor_fin = seccion[-1]
    linea_recta = np.linspace(valor_inicio, valor_fin, len(seccion))
    
    seccion_combinada = (seccion * (1.0 - ajuste_aplanado)) + (linea_recta * ajuste_aplanado)
    seccion_final = seccion_combinada + ajuste_offset
    datos_editados[inicio_idx:fin_idx+1] = seccion_final
    
    return datos_editados

# --- V23 MEJORA: Lógica de Aplanado Global ---
def generar_datos_con_ediciones_seccion(_datos_crudos_hoja, _config_por_dl, seed_value, ediciones_seccion):
    """
    Genera datos extrapolados aplicando:
    1. Pipeline base (incluye 'Escala de Picos')
    2. Aplanado Global (Estirar Hilo)
    3. Ediciones de Sección (incluye 'Aplanado de Sección')
    """
    datos_extrapolados = {}
    
    for dl_nombre, datos_originales in _datos_crudos_hoja.items():
        if dl_nombre in _config_por_dl:
            config_dl = _config_por_dl[dl_nombre]
            
            # 1. Aplicar el pipeline normal (que incluye "Escala de Picos")
            datos_procesados = aplicar_pipeline_a_columna(datos_originales, config_dl, seed_value)
            
            # --- ¡NUEVA SECCIÓN AQUÍ! ---
            # 2. Aplicar el APLANADO GLOBAL (Estirar Hilo)
            ajuste_aplanado_global = config_dl.get("aplanado_global", 0.0)
            if ajuste_aplanado_global > 0.0 and len(datos_procesados) > 1:
                valor_inicio = datos_procesados[0]
                valor_fin = datos_procesados[-1]
                linea_recta = np.linspace(valor_inicio, valor_fin, len(datos_procesados))
                
                # Mezclar la curva procesada con la línea recta
                datos_procesados = (datos_procesados * (1.0 - ajuste_aplanado_global)) + (linea_recta * ajuste_aplanado_global)
            # --- FIN DE LA NUEVA SECCIÓN ---
                
            # 3. Aplicar ediciones de sección (que pueden sobreescribir esto)
            if dl_nombre in ediciones_seccion:
                for edicion in ediciones_seccion[dl_nombre]:
                    datos_procesados = aplicar_edicion_seccion(
                        datos_procesados,
                        edicion['inicio_idx'],
                        edicion['fin_idx'],
                        edicion['ajuste_offset'],
                        edicion['ajuste_aplanado'] # Aplanado de Sección
                    )
            
            datos_extrapolados[dl_nombre] = datos_procesados
    
    return pd.DataFrame(dict([(k,pd.Series(v)) for k,v in datos_extrapolados.items()]))

# --- FUNCIONES DE MANEJO DE DATOS ---

@st.cache_data(show_spinner=False)
def leer_datos_crudos_excel(wb_bytes):
    """Lee TODOS los datos crudos del Excel y los almacena en un dict."""
    datos_crudos = {} # Estructura: { "hoja": { "DL": [datos] } }
    try:
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
        for hoja_nombre in wb.sheetnames:
            if any(ignorar in hoja_nombre.strip().upper() for ignorar in HOJAS_A_IGNORAR):
                continue
            
            ws = wb[hoja_nombre]
            datos_hoja = {}
            for col in ws.iter_cols(min_row=1):
                header_value = col[0].value
                if isinstance(header_value, str) and header_value.strip().upper().startswith("DL"):
                    valores = []
                    for cell in col[1:]:
                        if isinstance(cell.value, (int, float)):
                            valores.append(cell.value)
                    
                    if len(valores) > 20:
                        datos_hoja[header_value.strip()] = np.array(valores)
            
            if datos_hoja:
                datos_crudos[hoja_nombre] = datos_hoja
        return datos_crudos
    except Exception as e:
        st.error(f"Error al leer el archivo Excel: {e}")
        return None

def generar_configuracion_inicial(datos_crudos, config_base, seed_value):
    """Genera el dict de configuración inicial para CADA DL en TODAS las hojas."""
    random.seed(seed_value)
    np.random.seed(seed_value)
    
    config_hojas = {} # Estructura: { "hoja": { "DL": {config} } }

    for hoja_nombre, dls in datos_crudos.items():
        config_hoja_actual = {}
        for dl_nombre in dls.keys():
            config_hoja_actual[dl_nombre] = {
                "dl_nombre": dl_nombre,
                "variacion_percent": random.uniform(config_base["variacion_min"], config_base["variacion_max"]),
                "amplitud": config_base["amplitud"],
                "sigma": config_base["sigma"],
                "punto_pico_frac": config_base["punto_pico"],
                "offset_base": random.uniform(config_base["offset_min"], config_base["offset_max"]),
                "escala_picos": config_base["escala_picos"],
                "aplanado_global": config_base["aplanado_global"] # --- V23 MEJORA ---
            }
        config_hojas[hoja_nombre] = config_hoja_actual
            
    return config_hojas

@st.cache_data(show_spinner=False)
def generar_datos_extrapolados_df(_datos_crudos_hoja, _config_por_dl, seed_value):
    """Genera un DataFrame extrapolado (obsoleto, usar generar_datos_con_ediciones_seccion)"""
    # Esta función ya no es la principal, pero la mantenemos por si acaso.
    # La lógica real ahora está en generar_datos_con_ediciones_seccion
    datos_extrapolados = {}
    for dl_nombre, datos_originales in _datos_crudos_hoja.items():
        if dl_nombre in _config_por_dl:
            config_dl = _config_por_dl[dl_nombre]
            datos_extrapolados[dl_nombre] = aplicar_pipeline_a_columna(datos_originales, config_dl, seed_value)
    
    return pd.DataFrame(dict([(k,pd.Series(v)) for k,v in datos_extrapolados.items()]))


def dibujar_grafico_con_limites(df, titulo, limite_max=None, limite_min=None, seccion_seleccionada=None):
    """Crea un gráfico Altair con límites opcionales y sección seleccionada."""
    if df.empty:
        return st.warning(f"No se encontraron datos 'DL' para el gráfico: {titulo}")

    df_largo = df.reset_index().melt('index', var_name='Sensor', value_name='Valor')
    
    base = alt.Chart(df_largo).encode(
        x=alt.X('index', title='Índice de Tiempo'),
        y=alt.Y('Valor', title=titulo),
        color=alt.Color('Sensor', title="Sensores"),
        tooltip=['index', 'Sensor', 'Valor']
    ).properties(
        title=titulo
    )
    lineas = base.mark_line(point=False).interactive()
    grafico_final = lineas
    
    if seccion_seleccionada and 'inicio' in seccion_seleccionada and 'fin' in seccion_seleccionada:
        area_seleccionada = alt.Chart(pd.DataFrame({
            'x1': [seccion_seleccionada['inicio']],
            'x2': [seccion_seleccionada['fin']]
        })).mark_rect(
            opacity=0.3,
            color='yellow'
        ).encode(
            x='x1',
            x2='x2',
            y=alt.value(0),
            y2=alt.value(400)
        )
        grafico_final = grafico_final + area_seleccionada
    
    if limite_max is not None:
        linea_max = alt.Chart(pd.DataFrame({'y': [limite_max]})) \
            .mark_rule(color='red', strokeDash=[5, 2]) \
            .encode(y='y')
        grafico_final = grafico_final + linea_max
    if limite_min is not None:
        linea_min = alt.Chart(pd.DataFrame({'y': [limite_min]})) \
            .mark_rule(color='red', strokeDash=[5, 2]) \
            .encode(y='y')
        grafico_final = grafico_final + linea_min
        
    return st.altair_chart(grafico_final, use_container_width=True)


def descargar_excel_modificado(wb_bytes, config_hojas, seed_value, file_name, preset_archivo_base, ediciones_seccion=None):
    """Función final para procesar y descargar el Excel."""
    with st.spinner("Generando archivo Excel completo... Esto puede tardar unos segundos."):
        try:
            random.seed(seed_value)
            np.random.seed(seed_value)
            
            wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), keep_vba=True)
            
            aplicar_limite = any(nombre in preset_archivo_base for nombre in ARCHIVOS_CON_LIMITE)

            for hoja_nombre in wb.sheetnames:
                if hoja_nombre not in config_hojas:
                    continue 
                    
                config_a_usar = config_hojas[hoja_nombre]
                ws = wb[hoja_nombre]
                
                for col in ws.iter_cols(min_row=1):
                    header_cell = col[0]
                    header_value = str(header_cell.value).strip()
                    
                    if header_value in config_a_usar:
                        celdas_datos, valores_originales, tipos_originales = [], [], []
                        for cell in col[1:]:
                            if isinstance(cell.value, (int, float)) and cell.data_type != 'f':
                                celdas_datos.append(cell)
                                valores_originales.append(cell.value)
                                tipos_originales.append(type(cell.value))
                        
                        if len(valores_originales) < 20: continue
                        
                        datos_np = np.array(valores_originales)
                        config_dl = config_a_usar[header_value]
                        
                        # --- V23 MEJORA: Lógica de generación movida a 1 sola función ---
                        # Usamos la misma función que para los gráficos para consistencia
                        datos_finales = aplicar_pipeline_a_columna(datos_np, config_dl, seed_value)
                        
                        ajuste_aplanado_global = config_dl.get("aplanado_global", 0.0)
                        if ajuste_aplanado_global > 0.0 and len(datos_finales) > 1:
                            valor_inicio = datos_finales[0]
                            valor_fin = datos_finales[-1]
                            linea_recta = np.linspace(valor_inicio, valor_fin, len(datos_finales))
                            datos_finales = (datos_finales * (1.0 - ajuste_aplanado_global)) + (linea_recta * ajuste_aplanado_global)
                        
                        if ediciones_seccion and hoja_nombre in ediciones_seccion and header_value in ediciones_seccion[hoja_nombre]:
                            for edicion in ediciones_seccion[hoja_nombre][header_value]:
                                datos_finales = aplicar_edicion_seccion(
                                    datos_finales,
                                    edicion['inicio_idx'],
                                    edicion['fin_idx'],
                                    edicion['ajuste_offset'],
                                    edicion['ajuste_aplanado']
                                )
                        # --- Fin de la lógica ---

                        if aplicar_limite and ("%HR" not in hoja_nombre.upper() and "HUM" not in hoja_nombre.upper()):
                             np.clip(datos_finales, a_min=None, a_max=25.5, out=datos_finales)

                        for i, cell in enumerate(celdas_datos):
                            if i < len(datos_finales):
                                nuevo_valor = datos_finales[i]
                                tipo_original = tipos_originales[i]
                                if tipo_original == int: cell.value = int(round(nuevo_valor))
                                else:
                                    try: decimales = len(str(valores_originales[i]).split('.')[1])
                                    except: decimales = 2
                                    cell.value = round(nuevo_valor, decimales)

            with io.BytesIO() as f:
                wb.save(f)
                return f.getvalue()

        except Exception as e:
            st.error(f"Error al generar el archivo: {e}")
            logger.error(f"Error en descarga: {e}", exc_info=True)
            return None

# --- INTERFAZ DE STREAMLIT ---
st.set_page_config(layout="wide", page_title="Extrapolador Maestro V23")
st.title("Extrapolador Maestro V23 (Doble Aplanado) 🚀") # Título V23
st.info("Genera una extrapolación base, ajusta altura de picos (global) y 'estira' curvas (global o por sección).")

# --- BARRA LATERAL (CONTROLES GLOBALES) ---
st.sidebar.header("1. Carga de Archivo")
uploaded_file = st.sidebar.file_uploader("Cargar archivo .xlsm", type=["xlsm"])

# --- Inicializar estado de sesión para ediciones de sección ---
if 'ediciones_seccion' not in st.session_state:
    st.session_state.ediciones_seccion = {}

# --- LÓGICA PRINCIPAL (V23) ---
if uploaded_file is not None:
    
    if st.session_state.get('file_name') != uploaded_file.name:
        st.session_state.datos_crudos = leer_datos_crudos_excel(uploaded_file.getvalue())
        st.session_state.file_name = uploaded_file.name
        st.session_state.original_file_bytes = uploaded_file.getvalue()
        if 'config_hojas' in st.session_state: del st.session_state.config_hojas
        st.session_state.ediciones_seccion = {}
        st.cache_data.clear()

    if not st.session_state.datos_crudos:
        st.error("No se pudieron leer datos 'DL' válidos de este archivo. Revise el formato.")
    else:
        try:
            available_sheets = list(st.session_state.datos_crudos.keys())
            st.sidebar.header("2. Hoja de Trabajo")
            
            hoja_seleccionada = st.sidebar.selectbox(
                "Seleccionar Hoja para Visualizar y Editar", 
                available_sheets,
                index=0,
                key="hoja_seleccionada_principal"
            )

            st.sidebar.header("3. Generación Inicial")
            seed_value = st.sidebar.number_input("Versión (Semilla Aleatoria)", value=1, min_value=1, step=1, key="seed_value")
            preset_name = st.sidebar.selectbox("Seleccionar Preset de Prueba:", options=list(CONFIGURACION_BASE.keys()), key="preset_name")
            
            if st.sidebar.button("Generar/Reiniciar Extrapolación Base", type="primary"):
                with st.spinner("Generando extrapolación base..."):
                    config_base = CONFIGURACION_BASE[preset_name]
                    st.session_state.config_hojas = generar_configuracion_inicial(st.session_state.datos_crudos, config_base, seed_value)
                    st.session_state.last_seed = seed_value
                    st.session_state.last_preset = preset_name
                    st.session_state.ediciones_seccion = {}
                    st.toast(f"Generada Versión {seed_value} con preset '{preset_name}'")
                    st.rerun()
            
            if 'config_hojas' not in st.session_state or \
               st.session_state.get('last_seed') != seed_value or \
               st.session_state.get('last_preset') != preset_name:
                st.sidebar.warning("Haz clic en 'Generar/Reiniciar Base' para aplicar la nueva Versión o Preset.")

            if 'config_hojas' in st.session_state:
                
                st.sidebar.header("4. Editor Individual (Ajuste Global)")
                
                dl_names = list(st.session_state.datos_crudos[hoja_seleccionada].keys())
                dl_names_con_opcion_global = ["Aplicar a TODAS"] + dl_names
                dl_seleccionado = st.sidebar.selectbox("Curva a Editar:", dl_names_con_opcion_global)

                dl_para_config = dl_names[0] if dl_seleccionado == "Aplicar a TODAS" else dl_seleccionado
                
                if hoja_seleccionada not in st.session_state.config_hojas or dl_para_config not in st.session_state.config_hojas[hoja_seleccionada]:
                    st.sidebar.error(f"Error: No se encontró config para {hoja_seleccionada}. Intenta 'Generar Base' de nuevo.")
                else:
                    config_actual = st.session_state.config_hojas[hoja_seleccionada][dl_para_config]

                    # Slider de Escala de Picos (Global)
                    escala_picos = st.sidebar.slider(
                        "Escala de Picos (Global)", 0.0, 1.0, config_actual["escala_picos"], 0.1, 
                        help="Controla la altura de los picos. 1.0 = 100% original. 0.0 = 100% plano (picos eliminados).", 
                        key=f"escala_{dl_seleccionado}_{hoja_seleccionada}"
                    )
                    
                    # --- V23 MEJORA: Slider de Aplanado Global ---
                    aplanado_global = st.sidebar.slider(
                        "Aplanado Global (Estirar Hilo)", 0.0, 1.0, config_actual["aplanado_global"], 0.05,
                        help="Aplanar la curva *completa*. 1.0 = línea recta. 0.0 = original.",
                        key=f"aplanado_global_{dl_seleccionado}_{hoja_seleccionada}"
                    )
                    
                    # Sliders para edición completa (se mantienen)
                    variacion_percent = st.sidebar.slider(
                        "Extrapolación (Pico %)", 0.0, 0.2, config_actual["variacion_percent"], 0.01, 
                        help="Qué tanto 'sube' la curva en el pico.", key=f"var_{dl_seleccionado}_{hoja_seleccionada}"
                    )
                    offset_base = st.sidebar.slider(
                        "Nivel Vertical (Offset)", -2.0, 1.0, config_actual["offset_base"], 0.1, 
                        help="Sube o baja la curva completa.", key=f"offset_{dl_seleccionado}_{hoja_seleccionada}"
                    )
                    amplitud = st.sidebar.slider(
                        "Nivel de 'Unicidad' (Deriva)", 0.0, 1.0, config_actual["amplitud"], 0.05, 
                        help="Controla la 'personalidad' de la curva. 0 = plana.", key=f"amp_{dl_seleccionado}_{hoja_seleccionada}"
                    )
                    sigma = st.sidebar.slider(
                        "Suavidad de 'Unicidad' (Ondas)", 3, 25, config_actual["sigma"], 1, 
                        help="Longitud de las 'ondas'. 3 = cortas. 20 = largas.", key=f"sigma_{dl_seleccionado}_{hoja_seleccionada}"
                    )

                    # --- Lógica de Actualización en Tiempo Real ---
                    dl_a_actualizar = dl_names if dl_seleccionado == "Aplicar a TODAS" else [dl_seleccionado]
                    
                    for dl in dl_a_actualizar:
                        st.session_state.config_hojas[hoja_seleccionada][dl]["escala_picos"] = escala_picos
                        st.session_state.config_hojas[hoja_seleccionada][dl]["aplanado_global"] = aplanado_global # --- V23 MEJORA ---
                        st.session_state.config_hojas[hoja_seleccionada][dl]["variacion_percent"] = variacion_percent
                        st.session_state.config_hojas[hoja_seleccionada][dl]["offset_base"] = offset_base
                        st.session_state.config_hojas[hoja_seleccionada][dl]["amplitud"] = amplitud
                        st.session_state.config_hojas[hoja_seleccionada][dl]["sigma"] = sigma

                    # --- Editor de Sección Específica (V22) ---
                    st.sidebar.header("5. Editor de Sección Específica")
                    
                    dl_para_edicion_seccion = st.sidebar.selectbox(
                        "Seleccionar DL para editar sección:", 
                        dl_names,
                        key="dl_edicion_seccion"
                    )
                    
                    longitud_datos = len(st.session_state.datos_crudos[hoja_seleccionada][dl_para_edicion_seccion])
                    
                    col1, col2 = st.sidebar.columns(2)
                    with col1:
                        inicio_seccion = st.slider(
                            "Inicio sección", 
                            0, longitud_datos-1, 
                            max(0, longitud_datos//3),
                            key="inicio_seccion"
                        )
                    with col2:
                        fin_seccion = st.slider(
                            "Fin sección", 
                            0, longitud_datos-1, 
                            min(longitud_datos-1, longitud_datos*2//3),
                            key="fin_seccion"
                        )
                    
                    ajuste_offset_seccion = st.sidebar.slider(
                        "Ajuste Vertical (Offset)", -2.0, 2.0, 0.0, 0.1,
                        help="Sube o baja solo la sección seleccionada",
                        key="ajuste_offset_seccion"
                    )
                    
                    ajuste_aplanado_seccion = st.sidebar.slider(
                        "Aplanado (Estirar Hilo)", 0.0, 1.0, 0.0, 0.05,
                        help="Aplanar la curva en la sección. 1.0 = línea recta. 0.0 = original.",
                        key="ajuste_aplanado_seccion"
                    )
                    
                    col_apply, col_clear = st.sidebar.columns(2)
                    with col_apply:
                        if st.button("Aplicar a Sección", type="secondary"):
                            if hoja_seleccionada not in st.session_state.ediciones_seccion:
                                st.session_state.ediciones_seccion[hoja_seleccionada] = {}
                            if dl_para_edicion_seccion not in st.session_state.ediciones_seccion[hoja_seleccionada]:
                                st.session_state.ediciones_seccion[hoja_seleccionada][dl_para_edicion_seccion] = []
                            
                            nueva_edicion = {
                                'inicio_idx': inicio_seccion,
                                'fin_idx': fin_seccion,
                                'ajuste_offset': ajuste_offset_seccion,
                                'ajuste_aplanado': ajuste_aplanado_seccion 
                            }
                            st.session_state.ediciones_seccion[hoja_seleccionada][dl_para_edicion_seccion].append(nueva_edicion)
                            st.toast(f"Edición aplicada a {dl_para_edicion_seccion} en índices {inicio_seccion}-{fin_seccion}")
                            st.rerun()
                    
                    with col_clear:
                        if st.button("Limpiar Ediciones", type="secondary"):
                            if hoja_seleccionada in st.session_state.ediciones_seccion:
                                if dl_para_edicion_seccion in st.session_state.ediciones_seccion[hoja_seleccionada]:
                                    st.session_state.ediciones_seccion[hoja_seleccionada][dl_para_edicion_seccion] = []
                                    st.toast(f"Ediciones limpiadas para {dl_para_edicion_seccion}")
                                    st.rerun()
                    
                    if (hoja_seleccionada in st.session_state.ediciones_seccion and 
                        dl_para_edicion_seccion in st.session_state.ediciones_seccion[hoja_seleccionada] and
                        st.session_state.ediciones_seccion[hoja_seleccionada][dl_para_edicion_seccion]):
                        
                        st.sidebar.info("Ediciones activas:")
                        for i, edicion in enumerate(st.session_state.ediciones_seccion[hoja_seleccionada][dl_para_edicion_seccion]):
                            st.sidebar.write(f"{i+1}. Índices {edicion['inicio_idx']}-{edicion['fin_idx']}: "
                                           f"Offset={edicion['ajuste_offset']:.2f}, "
                                           f"Aplanado={edicion['ajuste_aplanado']:.2f}")

                    # --- Generar Gráficos ---
                    with st.spinner("Actualizando gráficos en tiempo real..."):
                        df_orig = pd.DataFrame(st.session_state.datos_crudos[hoja_seleccionada])
                        
                        df_ext = generar_datos_con_ediciones_seccion(
                            st.session_state.datos_crudos[hoja_seleccionada], 
                            st.session_state.config_hojas[hoja_seleccionada], 
                            seed_value,
                            st.session_state.ediciones_seccion.get(hoja_seleccionada, {})
                        )

                    # --- Área Principal (Gráficos) ---
                    st.header(f"Visualización (Hoja: {hoja_seleccionada})")
                    
                    limite_min, limite_max = None, None
                    titulo_grafico = "Original"
                    titulo_graf_ext = f"Extrapolado (Versión {seed_value})"
                    
                    if "%HR" not in hoja_seleccionada.upper() and "HUM" not in hoja_seleccionada.upper():
                        limite_min, limite_max = 15, 25
                        titulo_grafico = "Temperatura Original"
                        titulo_graf_ext = f"Temperatura Extrapolada (Versión {seed_value})"
                    else:
                        titulo_grafico = "Humedad Original"
                        titulo_graf_ext = f"Humedad Extrapolada (Versión {seed_value})"

                    seccion_seleccionada = None
                    if 'inicio_seccion' in st.session_state and 'fin_seccion' in st.session_state:
                        seccion_seleccionada = {
                            'inicio': st.session_state.inicio_seccion,
                            'fin': st.session_state.fin_seccion
                        }

                    col1, col2 = st.columns(2)
                    with col1:
                        dibujar_grafico_con_limites(df_orig, titulo_grafico, limite_max, limite_min)
                    with col2:
                        dibujar_grafico_con_limites(df_ext, titulo_graf_ext, limite_max, limite_min, seccion_seleccionada)
                    
                    # --- Botón de Descarga ---
                    st.sidebar.header("6. Descarga")
                    if st.sidebar.button(f"Generar y Descargar Excel (Versión {seed_value})"):
                        processed_bytes = descargar_excel_modificado(
                            st.session_state['original_file_bytes'], 
                            st.session_state.config_hojas, 
                            seed_value,
                            uploaded_file.name,
                            CONFIGURACION_BASE[preset_name]["archivo"],
                            st.session_state.ediciones_seccion
                        )
                        if processed_bytes:
                            st.sidebar.download_button(
                                label="¡Descarga Lista! (Haz clic aquí)",
                                data=processed_bytes,
                                file_name=f"extrapolado_v{seed_value}_{uploaded_file.name}",
                                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                                key="download_button"
                            )
                            st.sidebar.success("¡Archivo listo para descargar!")
                            st.rerun()

        except Exception as e:
            st.error(f"Error Crítico: {e}")
            logger.error(f"Error en Streamlit: {e}", exc_info=True)
            st.session_state.clear()

else:
    st.info("Cargue un archivo .xlsm para comenzar.")
    st.session_state.clear()
